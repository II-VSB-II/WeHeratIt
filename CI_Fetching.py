import requests
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

import sys
servicenow=sys.argv[1]
username=sys.argv[2]
password=sys.argv[3]
tablename=sys.argv[4]


url="https://"+servicenow+".service-now.com/api/now/cmdb/instance/"+tablename
headers = {"Content-Type": "application/json", "Accept": "application/json"}
response = requests.get(url, auth=(username,password), headers=headers)
cmdb = response.json()['result']

ws['A1'] = "NAME"
ws['B1'] = "HOSTNAME"
ws['C1'] = "CMDB ID"
i=2
for loop in cmdb:
    sys_id=loop["sys_id"].encode() 
    name=loop["name"].encode()
    data=requests.get(url+"/"+sys_id, auth=(username,password), headers=headers)
    data = data.json()['result']
    hostname=data["attributes"]["host_name"].encode()
    ws['A'+str(i)] = name
    ws['B'+str(i)] = hostname
    ws['C'+str(i)]=sys_id
    i=i+1
wb.save("CI_Data.xlsx")
